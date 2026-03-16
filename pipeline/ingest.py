"""Parse MOM wage data and SkillsFuture task-level data from Excel files."""

import json
import re
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz, process

from pipeline.models import MomOccupation, Occupation, SkillsFutureRole

RAW_DIR = Path("data/raw")
PROCESSED_DIR = Path("data/processed")

# SSOC major group labels
MAJOR_GROUP_LABELS = {
    1: ("managers", "Managers"),
    2: ("professionals", "Professionals"),
    3: ("associate_professionals", "Associate Professionals & Technicians"),
    4: ("clerical", "Clerical Support Workers"),
    5: ("service_sales", "Service & Sales Workers"),
    6: ("agricultural", "Skilled Agricultural & Fishery Workers"),
    7: ("craft", "Craft & Related Trades Workers"),
    8: ("plant_machine", "Plant & Machine Operators & Assemblers"),
    9: ("elementary", "Elementary Occupations"),
}


def slugify(title: str) -> str:
    """Convert occupation title to URL-safe slug."""
    s = title.lower().strip()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"[\s-]+", "-", s)
    return s.strip("-")


def parse_mom_wages(filepath: Path | None = None) -> list[MomOccupation]:
    """Parse MOM Occupational Wage Survey Excel (Table T4 — All Industries).

    Fixed layout (0-indexed columns):
      A(0): row number  B(1): SSOC 2020  C(2): Occupation
      D(3): Basic P25  E(4): Basic Median  F(5): Basic P75
      G(6): Gross P25  H(7): Gross Median  I(8): Gross P75

    Major group headers: SSOC = single digit, title = ALL CAPS (e.g. "1", "MANAGERS").
    Data rows start at Excel row 10 (0-indexed row 9 in openpyxl output).
    We use Gross Wage columns (G-I) as the primary wage measure.
    """
    if filepath is None:
        candidates = (
            list(RAW_DIR.glob("*table4*.*"))
            + list(RAW_DIR.glob("*Table4*.*"))
            + list(RAW_DIR.glob("*T4*.*"))
            + list(RAW_DIR.glob("*wage*.*"))
        )
        if not candidates:
            raise FileNotFoundError(
                f"No MOM wage file found in {RAW_DIR}/. "
                "Download mrsd_2024Wages_table4.xlsx from MOM and place in data/raw/."
            )
        filepath = candidates[0]

    print(f"Parsing MOM wages from: {filepath}")

    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Use sheet "T4" (All Industries)
    if "T4" in wb.sheetnames:
        ws = wb["T4"]
    else:
        ws = wb[wb.sheetnames[1]]  # First sheet after Contents

    occupations = []
    current_major_group = 0

    for row in ws.iter_rows(min_row=9, values_only=True):
        # row is a tuple of cell values
        ssoc_raw = str(row[1]).strip() if row[1] is not None else ""
        title_raw = str(row[2]).strip() if row[2] is not None else ""

        if not ssoc_raw or ssoc_raw == "None" or not title_raw or title_raw == "None":
            continue

        # Detect major group header rows (SSOC is single digit, title is ALL CAPS)
        ssoc_clean = re.sub(r"[^0-9]", "", ssoc_raw)
        if len(ssoc_clean) == 1 and title_raw == title_raw.upper() and not any(c.isdigit() for c in title_raw):
            current_major_group = int(ssoc_clean)
            continue

        if len(ssoc_clean) < 4:
            continue

        # Parse gross wage columns (G=index 6, H=index 7, I=index 8)
        gross_p25 = row[6] if len(row) > 6 else None
        gross_median = row[7] if len(row) > 7 else None
        gross_p75 = row[8] if len(row) > 8 else None

        if gross_median is None:
            continue

        try:
            monthly = int(float(gross_median))
        except (ValueError, TypeError):
            continue

        major_group = current_major_group if current_major_group > 0 else int(ssoc_clean[0])
        category, category_label = MAJOR_GROUP_LABELS.get(major_group, ("other", "Other"))

        try:
            p25 = int(float(gross_p25)) if gross_p25 is not None else None
        except (ValueError, TypeError):
            p25 = None
        try:
            p75 = int(float(gross_p75)) if gross_p75 is not None else None
        except (ValueError, TypeError):
            p75 = None

        occupations.append(
            MomOccupation(
                title=title_raw,
                slug=slugify(title_raw),
                ssoc_code=ssoc_clean,
                category=category,
                category_label=category_label,
                major_group=major_group,
                pay_monthly=monthly,
                pay_annual=monthly * 12,
                pay_p25=p25,
                pay_p75=p75,
            )
        )

    wb.close()
    print(f"  Parsed {len(occupations)} occupations")
    return occupations


def parse_skillsfuture(filepath: Path | None = None) -> list[SkillsFutureRole]:
    """Parse SkillsFuture Skills Framework combined database.

    The Excel has separate sheets for tasks and skills, each with one row per item:
    - 'Job Role_CWF_KT': Sector, Job Role, Critical Work Function, Key Tasks
    - 'Job Role_TSC': Sector, Job Role, TSC Title (technical skill competencies)
    - 'Job Role_Description': Sector, Job Role, Job Role Description
    """
    if filepath is None:
        # Prefer the full dataset over the template
        candidates = (
            list(RAW_DIR.glob("*skills-framework-dataset*.*"))
            + list(RAW_DIR.glob("*skillsfuture*.*"))
            + list(RAW_DIR.glob("*sfw*.*"))
        )
        if not candidates:
            raise FileNotFoundError(
                f"No SkillsFuture file found in {RAW_DIR}/. "
                "Download from SkillsFuture Interactive Skills Frameworks (requires Singpass)."
            )
        filepath = candidates[0]

    print(f"Parsing SkillsFuture from: {filepath}")
    sheets = pd.read_excel(filepath, sheet_name=None)
    print(f"  Found {len(sheets)} sheets: {list(sheets.keys())}")

    # Aggregate tasks per (sector, role) from Key Tasks sheet
    tasks_by_role: dict[tuple[str, str], list[str]] = {}
    if "Job Role_CWF_KT" in sheets:
        kt_df = sheets["Job Role_CWF_KT"]
        for _, row in kt_df.iterrows():
            sector = str(row.get("Sector", "")).strip()
            role = str(row.get("Job Role", "")).strip()
            task = str(row.get("Key Tasks", "")).strip()
            if not role or role == "nan" or not task or task == "nan":
                continue
            key = (sector, role)
            if key not in tasks_by_role:
                tasks_by_role[key] = []
            if task not in tasks_by_role[key]:
                tasks_by_role[key].append(task)

    # Aggregate skills per (sector, role) from TSC sheet
    # Column name varies: "TSC Title" (template) vs "TSC_CCS Title" (full dataset)
    skills_by_role: dict[tuple[str, str], list[str]] = {}
    tsc_sheet = "Job Role_TCS_CCS" if "Job Role_TCS_CCS" in sheets else "Job Role_TSC"
    if tsc_sheet in sheets:
        tsc_df = sheets[tsc_sheet]
        skill_col = "TSC_CCS Title" if "TSC_CCS Title" in tsc_df.columns else "TSC Title"
        # Only include TSC (technical skills), not CCS (core/soft skills)
        type_col = "TSC_CCS Type" if "TSC_CCS Type" in tsc_df.columns else None
        for _, row in tsc_df.iterrows():
            if type_col and str(row.get(type_col, "")).strip().lower() == "ccs":
                continue
            sector = str(row.get("Sector", "")).strip()
            role = str(row.get("Job Role", "")).strip()
            skill = str(row.get(skill_col, "")).strip()
            if not role or role == "nan" or not skill or skill == "nan":
                continue
            key = (sector, role)
            if key not in skills_by_role:
                skills_by_role[key] = []
            if skill not in skills_by_role[key]:
                skills_by_role[key].append(skill)

    # Build unique roles from all sheets
    all_keys: set[tuple[str, str]] = set()
    all_keys.update(tasks_by_role.keys())
    all_keys.update(skills_by_role.keys())

    # Also include roles from description sheet
    if "Job Role_Description" in sheets:
        desc_df = sheets["Job Role_Description"]
        for _, row in desc_df.iterrows():
            sector = str(row.get("Sector", "")).strip()
            role = str(row.get("Job Role", "")).strip()
            if role and role != "nan":
                all_keys.add((sector, role))

    roles: list[SkillsFutureRole] = []
    for sector, role_title in sorted(all_keys):
        key = (sector, role_title)
        roles.append(
            SkillsFutureRole(
                title=role_title,
                sector=sector,
                tasks=tasks_by_role.get(key, []),
                skills=skills_by_role.get(key, []),
            )
        )

    print(f"  Parsed {len(roles)} unique roles across {len({r.sector for r in roles})} sectors")
    return roles


def match_skillsfuture_to_mom(
    mom_occupations: list[MomOccupation],
    sf_roles: list[SkillsFutureRole],
) -> list[Occupation]:
    """Fuzzy-match SkillsFuture roles to MOM occupations by title.

    Two-pass matching:
    1. token_sort_ratio >= 70 (high confidence)
    2. WRatio >= 90 for remaining (catches partial matches without false positives)
    """
    sf_titles = [r.title for r in sf_roles]
    sf_by_title = {r.title: r for r in sf_roles}

    matched = 0
    occupations: list[Occupation] = []

    for mom in mom_occupations:
        sf_role = None

        # Pass 1: token_sort_ratio (strict word-level matching)
        best = process.extractOne(mom.title, sf_titles, scorer=fuzz.token_sort_ratio)
        if best and best[1] >= 70:
            sf_role = sf_by_title[best[0]]
        else:
            # Pass 2: WRatio (handles partial/substring matches) at high threshold
            best = process.extractOne(mom.title, sf_titles, scorer=fuzz.WRatio)
            if best and best[1] >= 90:
                sf_role = sf_by_title[best[0]]

        if sf_role:
            matched += 1

        occupations.append(
            Occupation(
                title=mom.title,
                slug=mom.slug,
                ssoc_code=mom.ssoc_code,
                category=mom.category,
                category_label=mom.category_label,
                major_group=mom.major_group,
                sector=sf_role.sector if sf_role else None,
                pay_monthly=mom.pay_monthly,
                pay_annual=mom.pay_annual,
                pay_p25=mom.pay_p25,
                pay_p75=mom.pay_p75,
                tasks=sf_role.tasks if sf_role else [],
                skills=sf_role.skills if sf_role else [],
                skillsfuture_funded=sf_role is not None,
            )
        )

    print(f"  Matched {matched}/{len(mom_occupations)} occupations to SkillsFuture roles")
    return occupations


def save_json(data: list, filepath: Path) -> None:
    """Save list of Pydantic models to JSON."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w") as f:
        json.dump([item.model_dump() if hasattr(item, "model_dump") else item for item in data], f, indent=2)
    print(f"  Saved {len(data)} records to {filepath}")


def main() -> None:
    """Run the ingestion pipeline."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # Step 1: Parse MOM wages
    print("\n=== Step 1: Parsing MOM wages ===")
    mom_occupations = parse_mom_wages()
    save_json(mom_occupations, PROCESSED_DIR / "mom_occupations.json")

    # Step 2: Parse SkillsFuture
    print("\n=== Step 2: Parsing SkillsFuture ===")
    try:
        sf_roles = parse_skillsfuture()
        save_json(sf_roles, PROCESSED_DIR / "skillsfuture_roles.json")
    except FileNotFoundError as e:
        print(f"  Warning: {e}")
        print("  Proceeding without SkillsFuture data (title-only scoring)")
        sf_roles = []

    # Step 3: Merge
    print("\n=== Step 3: Merging MOM + SkillsFuture ===")
    occupations = match_skillsfuture_to_mom(mom_occupations, sf_roles)
    save_json(occupations, PROCESSED_DIR / "occupations_enriched.json")

    print(f"\n✓ Ingestion complete: {len(occupations)} occupations")


if __name__ == "__main__":
    main()
